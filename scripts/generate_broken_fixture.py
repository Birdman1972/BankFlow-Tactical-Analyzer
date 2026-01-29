#!/usr/bin/env python3
"""
Generate BROKEN fixtures for testing Smart Repair.
Renames standard headers to non-standard ones.
"""
import argparse
import io
import os
import random
from generate_tc_fixtures import build_file_a, generate_shared_data, write_file, FILE_A_HEADERS

def build_broken_file_a(row_count: int, shared_data: list, seed: int = 42) -> bytes:
    # 1. Generate Valid Content First
    valid_bytes = build_file_a(row_count, shared_data, seed)
    
    # 2. Re-open and Hack Headers
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(valid_bytes))
    ws = wb.active
    
    # RENAME HEADERS to simulate "Bad Format"
    # "交易時間" -> "CustomTime"
    # "帳號" -> "UserID"
    header_row = ws[1]
    for cell in header_row:
        if cell.value == "交易時間":
            cell.value = "CustomTime"
        elif cell.value == "帳號":
            cell.value = "UserID"
            
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out-prefix", required=True)
    args = parser.parse_args()
    
    row_count = 50
    print(f"Generating BROKEN fixtures ({row_count} rows)...")
    
    shared = generate_shared_data(row_count)
    write_file(f"{args.out_prefix}_A.xlsx", build_broken_file_a(row_count, shared))
    # File B remains valid for this test (or could be broken too if needed)
    print("Done. Created broken_A.xlsx with headers 'CustomTime' and 'UserID'.")

if __name__ == "__main__":
    main()
