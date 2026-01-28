#!/usr/bin/env python3
"""
Generate rigorous BankFlow test fixtures with Traditional Chinese content.
"""
import argparse
import io
import math
import os
import random
from datetime import datetime, timedelta
from typing import List, Tuple

from openpyxl import Workbook

# --- Constants & Data Pools ---

FILE_A_HEADERS = [
    "交易時間",
    "帳號",
    "身分證字號",
    "戶名",
    "幣別",
    "交易序號",
    "摘要",
    "備註",
    "支出",
    "存入",
    "餘額",
    "對方帳號",
    "對方戶名",
]

FILE_B_HEADERS = [
    "登入時間",
    "帳號",
    "IP位址",
]

# Traditional Chinese Last Names
LAST_NAMES = ["陳", "林", "黃", "張", "李", "王", "吳", "劉", "蔡", "楊", "許", "鄭", "謝", "郭", "洪", "曾", "邱", "廖", "賴", "徐"]

# Traditional Chinese First Names (2 characters)
FIRST_NAMES = [
    "豪", "志明", "美玲", "淑芬", "俊傑", "雅婷", "冠宇", "宗翰", "家豪", "佳穎",
    "欣怡", "心怡", "志偉", "承恩", "宜君", "雅雯", "建宏", "怡君", "淑惠", "美惠",
    "佩君", "惠君", "嘉玲", "郁婷", "詩涵", "建志", "俊宏", "承翰", "冠志", "宇軒"
]

# Summaries (Transaction Types)
SUMMARIES = ["轉帳", "跨行提款", "現金存入", "薪資轉帳", "網購扣款", "一般消費", "代繳水費", "代繳電費", "信用卡款", "利息"]

# Remarks (Context)
REMARKS = [
    "網拍轉帳", "生活費", "還款", "聚餐", "買書", "繳房租", "不明", "測試備註", "蝦皮購物", "PChome",
    "UberEats", "Foodpanda", "自動扣繳", "手續費", "", "", "", ""  # Mix in some empty strings
]

PUBLIC_IPS = [
    "210.200.1.1", "220.135.10.10", "111.235.1.1", "59.124.1.1",  # TW IPs
    "8.8.8.8", "1.1.1.1", "140.112.1.1"
]

PRIVATE_IPS = [
    "192.168.1.10", "192.168.0.100", "10.0.0.5", "172.16.1.1"
]

# --- Generators ---

def random_name() -> str:
    return random.choice(LAST_NAMES) + random.choice(FIRST_NAMES)

def random_id(n: int) -> str:
    # Fake ID generator (not valid checksum, just format)
    first_char = chr(ord('A') + (n % 26))
    nums = f"{n % 1000000000:09d}"
    return f"{first_char}{nums}"

def build_file_a(row_count: int, seed: int = 42) -> bytes:
    random.seed(seed)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.append(FILE_A_HEADERS)

    base_time = datetime(2025, 4, 1, 9, 0, 0)
    balance = 500000.0

    # We will recycle a small set of accounts to ensure matches
    accounts = [f"{i:012d}" for i in range(min(100, row_count))] 

    for i in range(row_count):
        account = accounts[i % len(accounts)]
        
        # Add some time jitter
        time_offset = i * 60 + random.randint(0, 50) 
        timestamp = (base_time + timedelta(seconds=time_offset)).strftime("%Y/%m/%d %H:%M:%S")

        is_income = random.choice([True, False])
        income = 0.0
        expense = 0.0
        summary = random.choice(SUMMARIES)
        remark = random.choice(REMARKS)

        if is_income:
            income = float(random.randint(1, 500) * 100)
            balance += income
            if "扣" in summary or "費" in summary: # Fix semantic mismatch if any
                summary = "現金存入"
        else:
            expense = float(random.randint(1, 200) * 100)
            balance -= expense
            if "存" in summary or "薪" in summary:
                summary = "一般消費"

        # Generate Counterparty
        cp_account = ""
        cp_name = ""
        if summary in ["轉帳", "跨行提款", "薪資轉帳"]:
            cp_account = f"{random.randint(0, 999999):012d}"
            cp_name = random_name()

        row = [
            timestamp,          # 0 交易時間
            account,            # 1 帳號
            random_id(i),       # 2 身分證字號
            random_name(),      # 3 戶名 (Usually owner name, but let's vary it for test context)
            "TWD",              # 4 幣別
            f"TXN{i:08d}",      # 5 交易序號
            summary,            # 6 摘要
            remark,             # 7 備註
            expense if expense > 0 else "", # 8 支出
            income if income > 0 else "",   # 9 存入
            round(balance, 2),  # 10 餘額
            cp_account,         # 11 對方帳號
            cp_name             # 12 對方戶名
        ]
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def build_file_b(row_count: int, file_a_rows: int, seed: int = 42) -> bytes:
    random.seed(seed)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.append(FILE_B_HEADERS)

    # Reconstruct the same time/accounts to ensure matches
    base_time = datetime(2025, 4, 1, 9, 0, 0)
    accounts = [f"{i:012d}" for i in range(min(100, file_a_rows))]

    # We want to generate login records. 
    # Some should match exactly (approx time), some shouldn't.
    
    current_row = 0
    while current_row < row_count:
        # Pick an account
        idx = current_row % len(accounts)
        account = accounts[idx]
        
        # Corresponds roughly to the i-th transaction time for this account if we matched 1-to-1 linearly,
        # but let's just create a stream of logins relative to base time.
        
        # Create a "Match" opportunity every 3rd record
        if current_row < file_a_rows and current_row % 3 == 0:
            # Re-calculate File A time for this row index to create a match
            time_offset_a = current_row * 60 + random.randint(0, 50)
            tx_time = base_time + timedelta(seconds=time_offset_a)
            
            # Login time = Tx Time +/- deviation
            # Window is -1s to +2s. 
            # Let's hit it: Tx Time - 0s
            login_time = tx_time - timedelta(seconds=0)
            
            ip = random.choice(PUBLIC_IPS)
            
            ws.append([
                login_time.strftime("%Y/%m/%d %H:%M:%S"),
                account,
                ip
            ])
            current_row += 1
            
            # Occasional Multiple Match (2 logins in window)
            if random.random() < 0.1 and current_row < row_count:
                login_time_2 = tx_time + timedelta(seconds=1)
                ws.append([
                    login_time_2.strftime("%Y/%m/%d %H:%M:%S"),
                    account,
                    random.choice(PRIVATE_IPS) # Multi-IP scenario
                ])
                current_row += 1

        else:
            # Random noise login
            # Time moves forward
            global_offset = current_row * 50
            login_time = base_time + timedelta(seconds=global_offset)
            ws.append([
                login_time.strftime("%Y/%m/%d %H:%M:%S"),
                account,
                random.choice(PUBLIC_IPS + PRIVATE_IPS)
            ])
            current_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def write_file(path: str, data: bytes) -> None:
    path_obj = os.path.dirname(path)
    if path_obj:
        os.makedirs(path_obj, exist_ok=True)
    with open(path, "wb") as f:
        f.write(data)

def main() -> None:
    parser = argparse.ArgumentParser(description="Generate Traditional Chinese BankFlow Fixtures")
    parser.add_argument("--size", choices=["small", "large"], required=True, help="Dataset size")
    parser.add_argument("--out-prefix", required=True, help="Output file prefix (e.g. tests/fixtures/tc_small)")
    
    args = parser.parse_args()
    
    if args.size == "small":
        rows_a = 200
        rows_b = 300
    else:
        rows_a = 5000
        rows_b = 8000
    
    print(f"Generating {args.size} fixtures: A={rows_a}, B={rows_b} rows...")
    
    data_a = build_file_a(rows_a)
    out_a = f"{args.out_prefix}_A.xlsx"
    write_file(out_a, data_a)
    print(f"Created {out_a}")

    data_b = build_file_b(rows_b, rows_a)
    out_b = f"{args.out_prefix}_B.xlsx"
    write_file(out_b, data_b)
    print(f"Created {out_b}")

if __name__ == "__main__":
    main()
