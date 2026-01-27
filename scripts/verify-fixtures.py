#!/usr/bin/env python3

"""Verify canonical BankFlow XLSX fixtures.

This script validates that File A / File B fixtures follow the canonical schema
and that they contain enough data to exercise key features:

- income/expense split
- IP matching window ([-1s, +2s])
- counterparty extraction (File A col L)

It is intentionally schema/invariant-based (not exact output golden files).
"""

from __future__ import annotations

import argparse
import ipaddress
import sys
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable, Optional


try:
    from openpyxl import load_workbook
except Exception as e:  # pragma: no cover
    print("ERROR: openpyxl is required to run this script.", file=sys.stderr)
    print("Install: python3 -m pip install openpyxl", file=sys.stderr)
    raise


CANONICAL_FILE_A_HEADERS = [
    "交易時間",
    "帳號",
    "身分證字號",
    "戶名",
    "幣別",
    "交易序號",
    "摘要",
    "備註",
    "支出",
    "存入",
    "餘額",
    "對方帳號",
    "對方戶名",
]

CANONICAL_FILE_B_HEADERS = [
    "登入時間",
    "帳號",
    "IP位址",
]


def _parse_timestamp(value: object) -> datetime:
    if isinstance(value, datetime):
        return value
    if value is None or value == "":
        raise ValueError("empty timestamp")

    s = str(value).strip()
    for fmt in (
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d %H:%M",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f"unsupported timestamp format: {s!r}")


def _to_float(value: object) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    return float(s)


def _is_valid_ip(value: object) -> bool:
    if value is None or value == "":
        return False
    try:
        ipaddress.ip_address(str(value).strip())
        return True
    except ValueError:
        return False


def _read_headers(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = ["" if v is None else str(v).strip() for v in row]
    return headers


def _iter_rows(path: Path, max_rows: Optional[int]) -> Iterable[tuple[object, ...]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        yield row
        if max_rows is not None and (i + 1) >= max_rows:
            break


@dataclass(frozen=True)
class TxSample:
    timestamp: datetime
    account: str
    counterparty_account: str


@dataclass(frozen=True)
class IpSample:
    timestamp: datetime
    ip: str


def verify_file_a(path: Path, sample_rows: int) -> list[TxSample]:
    headers = _read_headers(path)
    if headers[: len(CANONICAL_FILE_A_HEADERS)] != CANONICAL_FILE_A_HEADERS:
        raise AssertionError(
            f"File A headers mismatch in {path}.\n"
            f"Expected: {CANONICAL_FILE_A_HEADERS}\n"
            f"Got:      {headers}"
        )

    samples: list[TxSample] = []
    income_rows = 0
    expense_rows = 0
    counterparty_nonempty = 0

    for row in _iter_rows(path, max_rows=sample_rows):
        ts = _parse_timestamp(row[0])
        account = str(row[1]).strip()
        if not account:
            continue

        expense = _to_float(row[8])
        income = _to_float(row[9])
        if expense is not None and expense > 0:
            expense_rows += 1
        if income is not None and income > 0:
            income_rows += 1

        cp = "" if row[11] is None else str(row[11]).strip()
        if cp:
            counterparty_nonempty += 1

        samples.append(TxSample(timestamp=ts, account=account, counterparty_account=cp))

    if not samples:
        raise AssertionError(f"File A has no data rows: {path}")
    if income_rows == 0 or expense_rows == 0:
        raise AssertionError(
            f"File A does not exercise split logic (need both income and expense rows): {path}"
        )
    if counterparty_nonempty == 0:
        raise AssertionError(
            f"File A does not contain any counterparty accounts (col L): {path}"
        )

    return samples


def verify_file_b(path: Path, sample_rows: int) -> dict[str, list[IpSample]]:
    headers = _read_headers(path)
    if headers[: len(CANONICAL_FILE_B_HEADERS)] != CANONICAL_FILE_B_HEADERS:
        raise AssertionError(
            f"File B headers mismatch in {path}.\n"
            f"Expected: {CANONICAL_FILE_B_HEADERS}\n"
            f"Got:      {headers}"
        )

    records: dict[str, list[IpSample]] = {}
    valid_ip_rows = 0

    for row in _iter_rows(path, max_rows=sample_rows):
        ts = _parse_timestamp(row[0])
        account = str(row[1]).strip()
        ip = "" if row[2] is None else str(row[2]).strip()
        if not account:
            continue
        if _is_valid_ip(ip):
            valid_ip_rows += 1

        records.setdefault(account, []).append(IpSample(timestamp=ts, ip=ip))

    if not records:
        raise AssertionError(f"File B has no data rows: {path}")
    if valid_ip_rows == 0:
        raise AssertionError(f"File B contains no valid IP addresses: {path}")

    for items in records.values():
        items.sort(key=lambda x: x.timestamp)
    return records


def verify_ip_window(
    tx_samples: list[TxSample], ip_by_account: dict[str, list[IpSample]]
) -> None:
    window_before = timedelta(seconds=1)
    window_after = timedelta(seconds=2)

    matched = 0
    multi_ip = 0

    for tx in tx_samples:
        items = ip_by_account.get(tx.account)
        if not items:
            continue

        start = tx.timestamp - window_before
        end = tx.timestamp + window_after
        ips = {
            it.ip
            for it in items
            if start <= it.timestamp <= end and _is_valid_ip(it.ip)
        }
        if ips:
            matched += 1
        if len(ips) >= 2:
            multi_ip += 1

    if matched == 0:
        raise AssertionError(
            "No transactions had any matching IP records within [-1s, +2s]"
        )

    # We don't enforce exact ratios, only that the fixture exercises both paths.
    if multi_ip == 0:
        raise AssertionError(
            "No transactions produced multi-IP matches (need at least one)"
        )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Verify BankFlow canonical XLSX fixtures"
    )
    parser.add_argument(
        "--sample-rows",
        type=int,
        default=5000,
        help="Max data rows to sample from each file (default: 5000)",
    )
    parser.add_argument(
        "--file-a",
        type=Path,
        default=Path("tests/fixtures/test_transactions_1000.xlsx"),
        help="Path to File A fixture",
    )
    parser.add_argument(
        "--file-b",
        type=Path,
        default=Path("tests/fixtures/test_ip_records_1000.xlsx"),
        help="Path to File B fixture",
    )
    args = parser.parse_args()

    file_a = args.file_a
    file_b = args.file_b
    if not file_a.exists():
        print(f"ERROR: File A not found: {file_a}", file=sys.stderr)
        return 2
    if not file_b.exists():
        print(f"ERROR: File B not found: {file_b}", file=sys.stderr)
        return 2

    tx_samples = verify_file_a(file_a, sample_rows=args.sample_rows)
    ip_by_account = verify_file_b(file_b, sample_rows=args.sample_rows)
    verify_ip_window(tx_samples, ip_by_account)

    print("OK: fixtures verified")
    print(f"- File A: {file_a}")
    print(f"- File B: {file_b}")
    print(f"- sampled rows (max): {args.sample_rows}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
