#!/usr/bin/env python3
import io
import math
import os
from datetime import datetime, timedelta

from openpyxl import Workbook

FILE_A_HEADERS = [
    "交易序號",
    "帳號",
    "客戶姓名",
    "交易時間",
    "交易類型",
    "身分證/統編",
    "交易摘要",
    "交易後餘額",
    "支出金額",
    "存入金額",
    "幣別",
    "聯絡電話",
    "通訊地址",
    "備註",
]

FILE_B_HEADERS = [
    "登入序號",
    "帳號",
    "登入時間",
    "IP位址",
    "裝置資訊",
    "登入地區",
]

PUBLIC_IPS = [
    "8.8.8.8",
    "1.1.1.1",
    "208.67.222.222",
    "9.9.9.9",
]

PRIVATE_IPS = [
    "10.0.0.1",
    "192.168.1.1",
    "172.16.0.1",
]


def build_file_a(row_count: int) -> bytes:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.append(FILE_A_HEADERS)

    base_time = datetime(2024, 1, 15, 10, 30, 0)
    balance = 100000.0

    for i in range(row_count):
        account = f"ACC{i % 100000:05d}"
        timestamp = (base_time + timedelta(seconds=i)).strftime("%Y-%m-%d %H:%M:%S")
        income = 0.0
        expense = 0.0
        if i % 2 == 0:
            income = float((i % 50 + 1) * 100)
            balance += income
            tx_type = "入帳"
        else:
            expense = float((i % 40 + 1) * 80)
            balance -= expense
            tx_type = "出帳"

        row = [
            i + 1,
            account,
            f"客戶{i % 500:03d}",
            timestamp,
            tx_type,
            f"A{i % 999999999:09d}",
            "轉帳",
            round(balance, 2),
            expense,
            income,
            "TWD",
            f"09{i % 100000000:08d}",
            f"台北市中正區{i % 100}號",
            "測試資料",
        ]
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_file_b(row_count: int) -> bytes:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.append(FILE_B_HEADERS)

    base_time = datetime(2024, 1, 15, 10, 30, 1)
    for i in range(row_count):
        account = f"ACC{i % 100000:05d}"
        timestamp = (base_time + timedelta(seconds=i)).strftime("%Y-%m-%d %H:%M:%S")
        ip = PUBLIC_IPS[i % len(PUBLIC_IPS)]
        row = [
            i + 1,
            account,
            timestamp,
            ip,
            "Chrome",
            "台灣",
        ]
        ws.append(row)

        if i % 500 == 0:
            multi_ip = PRIVATE_IPS[(i // 500) % len(PRIVATE_IPS)]
            row2 = [
                f"{i + 1}-2",
                account,
                (base_time + timedelta(seconds=i + 1)).strftime("%Y-%m-%d %H:%M:%S"),
                multi_ip,
                "Firefox",
                "台灣",
            ]
            ws.append(row2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def estimate_rows(target_mb: int, builder) -> int:
    sample_rows = 1000
    sample_bytes = builder(sample_rows)
    bytes_per_row = max(1, len(sample_bytes) / sample_rows)
    target_bytes = target_mb * 1024 * 1024
    return int(math.ceil(target_bytes / bytes_per_row))


def write_file(path: str, data: bytes) -> None:
    with open(path, "wb") as f:
        f.write(data)


def generate(target_mb: int, out_a: str, out_b: str) -> None:
    rows_a = estimate_rows(target_mb, build_file_a)
    rows_b = estimate_rows(target_mb, build_file_b)

    data_a = build_file_a(rows_a)
    data_b = build_file_b(rows_b)

    write_file(out_a, data_a)
    write_file(out_b, data_b)

    print(f"Generated {out_a}: {os.path.getsize(out_a) / (1024*1024):.2f} MB, rows={rows_a}")
    print(f"Generated {out_b}: {os.path.getsize(out_b) / (1024*1024):.2f} MB, rows={rows_b}")


def main() -> None:
    generate(20, "tmp_test_FileA_20mb.xlsx", "tmp_test_FileB_20mb.xlsx")
    generate(50, "tmp_test_FileA_50mb.xlsx", "tmp_test_FileB_50mb.xlsx")


if __name__ == "__main__":
    main()
